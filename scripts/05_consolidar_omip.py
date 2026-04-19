import os
import re
import shutil
import datetime as dt

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) if "__file__" in globals() else os.getcwd()
TEMPLATE_PATH = os.path.join(BASE_DIR, "inputs", "OMIP_Template.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "data", "OMIP_actualizado.xlsx")
CSV_ES_PATH = os.path.join(BASE_DIR, "data", "omip_futuros_es.csv")
CSV_PT_PATH = os.path.join(BASE_DIR, "data", "omip_futuros_pt.csv")

MONTH_ALIASES = {
    "jan": "Jan",
    "feb": "Feb",
    "mar": "Mrz",
    "mrz": "Mrz",
    "apr": "Apr",
    "may": "Mai",
    "mai": "Mai",
    "jun": "Jun",
    "jul": "Jul",
    "aug": "Aug",
    "ago": "Aug",
    "sep": "Sep",
    "oct": "Okt",
    "okt": "Okt",
    "nov": "Nov",
    "dec": "Dez",
    "dez": "Dez",
    "dic": "Dez",
}

GERMAN_MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mrz": 3,
    "Apr": 4,
    "Mai": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Dez": 12,
}


def clean_text(value):
    if value is None:
        return ""
    s = str(value).replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_header(value: str) -> str:
    s = clean_text(value)
    if not s:
        return ""

    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()

    m = re.fullmatch(r"q([1-4])\s*(\d{2})", s, flags=re.IGNORECASE)
    if m:
        return f"Q{m.group(1)} {m.group(2)}"

    m = re.fullmatch(r"(?:cal|yr|y)\s*(\d{2})", s, flags=re.IGNORECASE)
    if m:
        return f"Cal {m.group(1)}"

    m = re.fullmatch(r"([A-Za-z]+)\s*(\d{2})", s)
    if m:
        month_raw = m.group(1).lower()
        yy = m.group(2)
        month = MONTH_ALIASES.get(month_raw)
        if month:
            return f"{month} {yy}"

    return s


def parse_sheet_date(value) -> pd.Timestamp | None:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date, pd.Timestamp)):
        return pd.Timestamp(value).normalize()

    s = clean_text(value)
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return pd.Timestamp(dt.datetime.strptime(s, fmt).date())
        except ValueError:
            continue
    return None


def header_to_expiry(header: str) -> pd.Timestamp | None:
    if not header:
        return None

    s = normalize_header(header)

    m = re.fullmatch(r"(Jan|Feb|Mrz|Apr|Mai|Jun|Jul|Aug|Sep|Okt|Nov|Dez) (\d{2})", s)
    if m:
        month = GERMAN_MONTHS[m.group(1)]
        year = 2000 + int(m.group(2))
        return pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)

    q = re.fullmatch(r"Q([1-4]) (\d{2})", s)
    if q:
        quarter = int(q.group(1))
        year = 2000 + int(q.group(2))
        month = quarter * 3
        return pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)

    y = re.fullmatch(r"Cal (\d{2})", s)
    if y:
        year = 2000 + int(y.group(1))
        return pd.Timestamp(year=year, month=12, day=31)

    return None


def load_csv(path: str) -> pd.DataFrame | None:
    if not os.path.exists(path):
        return None

    df = pd.read_csv(path)
    if df.empty:
        return None

    required = {"TRADE_DATE", "EXCEL_HEADER", "PRICE_USED"}
    if not required.issubset(df.columns):
        return None

    df["TRADE_DATE"] = pd.to_datetime(df["TRADE_DATE"]).dt.normalize()
    df["EXCEL_HEADER_NORM"] = df["EXCEL_HEADER"].apply(normalize_header)
    df = df[df["EXCEL_HEADER_NORM"].notna() & (df["EXCEL_HEADER_NORM"] != "") & df["PRICE_USED"].notna()].copy()
    return df


def build_maps(ws):
    header_to_col = {}
    header_raw = {}

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is not None:
            raw = clean_text(header)
            norm = normalize_header(raw)
            if norm:
                header_to_col[norm] = col
                header_raw[norm] = raw

    date_to_row = {}
    for row in range(2, ws.max_row + 1):
        date_value = parse_sheet_date(ws.cell(row=row, column=1).value)
        if date_value is not None:
            date_to_row[date_value] = row

    return header_to_col, header_raw, date_to_row


def actualizar_hoja(ws, df: pd.DataFrame, limpiar_vencidos: bool = True, debug_name: str = ""):
    header_to_col, header_raw, date_to_row = build_maps(ws)

    escritos = 0
    missing_headers = set()
    missing_dates = 0

    for row in df.itertuples(index=False):
        header = row.EXCEL_HEADER_NORM
        trade_date = pd.Timestamp(row.TRADE_DATE).normalize()

        if header not in header_to_col:
            missing_headers.add(header)
            continue

        if trade_date not in date_to_row:
            missing_dates += 1
            continue

        col = header_to_col[header]
        row_idx = date_to_row[trade_date]
        ws.cell(row=row_idx, column=col, value=float(row.PRICE_USED))
        escritos += 1

    limpiados = 0
    if limpiar_vencidos:
        sorted_dates = sorted(date_to_row.items(), key=lambda x: x[0])
        for header_norm, col in header_to_col.items():
            if col == 1:
                continue
            expiry = header_to_expiry(header_norm)
            if expiry is None:
                continue
            for current_date, row_idx in sorted_dates:
                if current_date > expiry:
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value is not None:
                        cell.value = None
                        limpiados += 1

    if missing_headers:
        print(f"{debug_name}: headers no encontrados ({len(missing_headers)}): {sorted(list(missing_headers))[:20]}")
    if missing_dates:
        print(f"{debug_name}: fechas no encontradas: {missing_dates}")

    return escritos, limpiados


def main():
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"No existe: {TEMPLATE_PATH}")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    shutil.copyfile(TEMPLATE_PATH, OUTPUT_PATH)

    wb = load_workbook(OUTPUT_PATH)

    df_es = load_csv(CSV_ES_PATH)
    df_pt = load_csv(CSV_PT_PATH)

    if df_es is not None and "Spain OMIP" in wb.sheetnames:
        escritos, limpiados = actualizar_hoja(wb["Spain OMIP"], df_es, limpiar_vencidos=True, debug_name="Spain OMIP")
        print(f"Spain OMIP: escritos={escritos}, limpiados={limpiados}")
    else:
        print("Spain OMIP: sin CSV nuevo. Se conserva la plantilla.")

    if df_pt is not None and "Portugal OMIP" in wb.sheetnames:
        escritos, limpiados = actualizar_hoja(wb["Portugal OMIP"], df_pt, limpiar_vencidos=True, debug_name="Portugal OMIP")
        print(f"Portugal OMIP: escritos={escritos}, limpiados={limpiados}")
    else:
        print("Portugal OMIP: sin CSV nuevo. Se conserva la plantilla.")

    wb.save(OUTPUT_PATH)
    print(f"Archivo generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
